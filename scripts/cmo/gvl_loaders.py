"""GVL (DE neighbouring rights) bulk loaders — listen, KONU, sendemeldungen."""

from __future__ import annotations

import csv
import re
from pathlib import Path

import openpyxl
from pypdf import PdfReader

from loaders import pack_source, tokens_from

ISRC_RE = re.compile(r"\b[A-Z]{2}[A-Z0-9]{3}\d{7}\b")
RECORD_START_RE = re.compile(r"^(\d+)\s+(Tonträger|Videoclip)\s+(.+)$", re.MULTILINE)
DATE_RE = re.compile(r"\b\d{2}\.\d{2}\.\d{4}\b")

SENDEMELDUNG_CSV_FIELDS = (
    "gvl_id",
    "medientyp",
    "interpret",
    "isrc",
    "nutzungsdatum",
    "sender",
    "year",
    "pdf_source",
)

LISTEN_LABELS = {
    "GVL_Rightsholders_AddressUnknown_Artists.xlsx": "listen-artists: cím ismeretlen",
    "GVL_Rightsholders_DeceasedHeirsUnknown_Artists.xlsx": "listen-artists: örökös ismeretlen",
    "GVL_Rightsholders_DataIncomplete_Artists.xlsx": "listen-artists: hiányos adat",
    "GVL_Rightsholders_AddressUnknown_Producers.xlsx": "listen-producers: cím ismeretlen",
}


def _find_header_row(ws) -> tuple[int, tuple]:
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        if not row:
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        if cells[0].lower() == "name":
            return i, row
    raise ValueError("name column header not found")


def load_gvl_listen(base: Path) -> list[dict]:
    listen_dir = base / "listen"
    if not listen_dir.is_dir():
        return []

    records: list[dict] = []
    seen: set[str] = set()

    for path in sorted(listen_dir.glob("*.xlsx")):
        label = LISTEN_LABELS.get(path.name, f"listen: {path.stem}")
        is_producers = "Producers" in path.name
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header_row_num, header_row = _find_header_row(ws)
        headers = [str(c).strip().lower() if c else "" for c in header_row]

        name_idx = headers.index("name")
        vorname_idx = headers.index("vorname") if "vorname" in headers else None
        ort_idx = headers.index("ort") if "ort" in headers else None
        if is_producers and "ort/city" in headers:
            ort_idx = headers.index("ort/city")

        for row_num, row in enumerate(ws.iter_rows(min_row=header_row_num + 1, values_only=True), start=1):
            if not row:
                continue
            name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] else ""
            if not name:
                continue
            vorname = ""
            if vorname_idx is not None and vorname_idx < len(row) and row[vorname_idx]:
                vorname = str(row[vorname_idx]).strip()
            ort = ""
            if ort_idx is not None and ort_idx < len(row) and row[ort_idx]:
                ort = str(row[ort_idx]).strip()

            performer = name if is_producers else f"{vorname} {name}".strip() if vorname else name
            rec_id = f"{path.stem}:{row_num}"
            if rec_id in seen:
                continue
            seen.add(rec_id)

            remark = ort or None
            records.append(
                {
                    "id": rec_id,
                    "source": "de-gvl",
                    "title": label,
                    "identification": performer,
                    "performer": performer,
                    "remark": remark,
                    "gvlList": "listen-producers" if is_producers else "listen-artists",
                }
            )
        wb.close()

    return records


def _load_konu_file(path: Path, *, year: int) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return []

    medium = path.stem.split("_")[1] if "_" in path.stem else path.stem
    records: list[dict] = []
    seen: set[str] = set()

    for row_num, row in enumerate(rows_iter, start=2):
        if not row or len(row) < 9:
            continue
        prod_num = str(row[1]).strip() if row[1] is not None else ""
        if not prod_num or prod_num.lower() == "produktionsnummer - production number":
            continue
        title = str(row[2]).strip() if row[2] else ""
        additional = str(row[3]).strip() if row[3] else ""
        performer = str(row[5]).strip() if row[5] else ""
        composer = str(row[6]).strip() if row[6] else ""
        producer = str(row[7]).strip() if row[7] else ""
        isrc_raw = str(row[8]).strip() if row[8] else ""
        label = str(row[12]).strip() if len(row) > 12 and row[12] else ""

        full_title = title
        if additional:
            full_title = f"{title} ({additional})" if title else additional
        if not full_title and not performer:
            continue

        rec_id = f"{path.name}:{prod_num}"
        if rec_id in seen:
            continue
        seen.add(rec_id)

        ident_parts = [p for p in (performer, additional, composer, producer) if p]
        ident = " · ".join(ident_parts) if ident_parts else performer

        rec: dict = {
            "id": rec_id,
            "source": "de-gvl",
            "title": full_title or "(névtelen)",
            "identification": ident,
            "remark": f"KONU {medium} {year}",
            "gvlList": "produktionen",
            "gvlYear": year,
            "gvlMedium": medium,
        }
        if performer:
            rec["performer"] = performer
        if composer:
            rec["composer"] = composer
        if additional:
            rec["gvlRemix"] = additional
        if label:
            rec["label"] = label
        if isrc_raw:
            rec["isrc"] = isrc_raw.upper()
        records.append(rec)

    wb.close()
    return records


def load_gvl_produktionen(base: Path) -> list[dict]:
    records: list[dict] = []
    for sub in ("produktionen_2022", "produktionen_2023"):
        pdir = base / sub
        if not pdir.is_dir():
            continue
        year = int(sub.rsplit("_", 1)[-1])
        for path in sorted(pdir.glob("KONU_*.xlsx")):
            records.extend(_load_konu_file(path, year=year))
    return records


def _parse_pdf_block(block: str, *, year: int, medientyp_default: str) -> dict | None:
    row = _parse_pdf_block_row(block, year=year)
    if not row:
        return None
    return _sendemeldung_row_to_record(row)


def _parse_pdf_block_row(block: str, *, year: int) -> dict | None:
    block = block.strip()
    if not block:
        return None
    m = RECORD_START_RE.match(block)
    if not m:
        return None
    rec_id, medientyp, rest = m.group(1), m.group(2), m.group(3)
    isrcs = ISRC_RE.findall(block)
    if not isrcs:
        return None
    isrc = isrcs[0]

    pre_isrc = rest
    idx = rest.find(isrc)
    if idx >= 0:
        pre_isrc = rest[:idx].strip()
    pre_isrc = re.sub(r"\s+", " ", pre_isrc.replace("\n", " ")).strip()

    post = rest[idx + len(isrc) :] if idx >= 0 else ""
    post = re.sub(r"\s+", " ", post.replace("\n", " ")).strip()
    date_m = DATE_RE.search(post)
    date = date_m.group(0) if date_m else ""
    sender = post[date_m.end() :].strip() if date_m else post

    return {
        "gvl_id": rec_id,
        "medientyp": medientyp,
        "interpret": pre_isrc,
        "isrc": isrc,
        "nutzungsdatum": date,
        "sender": sender,
        "year": str(year),
        "pdf_source": "",
    }


def _sendemeldung_row_to_record(row: dict) -> dict:
    year = int(row["year"])
    medientyp = row["medientyp"]
    rec_id = row["gvl_id"]
    pre_isrc = row["interpret"]
    isrc = row["isrc"].upper()
    date = row.get("nutzungsdatum") or None
    sender = row.get("sender") or ""

    performer = pre_isrc
    if "," in pre_isrc[:80]:
        parts = pre_isrc.split(" ", 1)
        performer = parts[0] if parts else pre_isrc

    remark_bits = [f"Sendemeldung {year}", medientyp]
    if date:
        remark_bits.append(date)
    if sender:
        remark_bits.append(sender)

    return {
        "id": f"sendemeldung:{year}:{medientyp}:{rec_id}",
        "source": "de-gvl",
        "title": pre_isrc or "(névtelen)",
        "identification": pre_isrc or performer,
        "performer": performer or None,
        "remark": " · ".join(remark_bits),
        "isrc": isrc,
        "gvlList": "sendemeldungen",
        "gvlYear": year,
        "gvlMedium": medientyp,
    }


def _parse_sendemeldung_text(text: str, *, year: int) -> list[dict]:
    records: list[dict] = []
    seen: set[str] = set()
    parts = RECORD_START_RE.split(text)
    i = 1
    while i + 2 < len(parts):
        rec_id, medientyp, body = parts[i], parts[i + 1], parts[i + 2]
        block = f"{rec_id} {medientyp} {body}"
        row = _parse_pdf_block_row(block, year=year)
        if row:
            row["pdf_source"] = ""
            dedupe = f"{year}:{medientyp}:{row['gvl_id']}"
            if dedupe not in seen:
                seen.add(dedupe)
                records.append(row)
        i += 3
    return records


def _parse_sendemeldung_records(text: str, *, year: int) -> list[dict]:
    return [_sendemeldung_row_to_record(r) for r in _parse_sendemeldung_text(text, year=year)]


def sendemeldung_csv_path(csv_dir: Path, pdf_path: Path) -> Path:
    return csv_dir / f"{pdf_path.stem}.csv"


def _pdf_text_cache_path(cache_dir: Path, pdf_path: Path) -> Path:
    return cache_dir / f"{pdf_path.stem}.txt"


def _extract_pdf_text(pdf_path: Path, cache_dir: Path) -> str:
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = _pdf_text_cache_path(cache_dir, pdf_path)
    if cache_path.is_file() and cache_path.stat().st_mtime >= pdf_path.stat().st_mtime:
        return cache_path.read_text(encoding="utf-8", errors="replace")

    reader = PdfReader(str(pdf_path))
    chunks: list[str] = []
    for page in reader.pages:
        chunks.append(page.extract_text() or "")
    text = "\n".join(chunks)
    cache_path.write_text(text, encoding="utf-8")
    return text


def _sendemeldung_source_text(pdf_path: Path, *, text_cache_dir: Path) -> str:
    cache_path = _pdf_text_cache_path(text_cache_dir, pdf_path)
    if cache_path.is_file() and cache_path.stat().st_mtime >= pdf_path.stat().st_mtime:
        return cache_path.read_text(encoding="utf-8", errors="replace")
    return _extract_pdf_text(pdf_path, text_cache_dir)


def export_sendemeldung_csv(
    pdf_path: Path,
    csv_path: Path,
    *,
    year: int,
    text_cache_dir: Path,
    force: bool = False,
) -> int:
    if (
        not force
        and csv_path.is_file()
        and csv_path.stat().st_mtime >= pdf_path.stat().st_mtime
    ):
        with csv_path.open(encoding="utf-8", newline="") as f:
            return sum(1 for _ in csv.DictReader(f))

    print(f"  GVL PDF → CSV: {pdf_path.name} …", flush=True)
    text = _sendemeldung_source_text(pdf_path, text_cache_dir=text_cache_dir)
    rows = _parse_sendemeldung_text(text, year=year)
    for row in rows:
        row["pdf_source"] = pdf_path.name

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=SENDEMELDUNG_CSV_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"    → {len(rows):,} rows → {csv_path.name}", flush=True)
    return len(rows)


def load_sendemeldung_csv(csv_path: Path) -> list[dict]:
    records: list[dict] = []
    seen: set[str] = set()
    with csv_path.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            dedupe = f"{row['year']}:{row['medientyp']}:{row['gvl_id']}"
            if dedupe in seen:
                continue
            seen.add(dedupe)
            records.append(_sendemeldung_row_to_record(row))
    return records


def ensure_sendemeldungen_csv(
    base: Path,
    *,
    csv_dir: Path,
    text_cache_dir: Path,
    force: bool = False,
) -> list[Path]:
    send_dir = base / "sendemeldungen"
    if not send_dir.is_dir():
        return []

    csv_paths: list[Path] = []
    for pdf_path in sorted(send_dir.glob("GVL_Offene_Nutzungen_*.pdf")):
        year_m = re.search(r"_(\d{4})_", pdf_path.name)
        if not year_m:
            continue
        year = int(year_m.group(1))
        csv_path = sendemeldung_csv_path(csv_dir, pdf_path)
        export_sendemeldung_csv(
            pdf_path,
            csv_path,
            year=year,
            text_cache_dir=text_cache_dir,
            force=force,
        )
        csv_paths.append(csv_path)
    return csv_paths


def load_gvl_sendemeldungen(
    base: Path,
    *,
    csv_dir: Path,
    text_cache_dir: Path,
    force_csv: bool = False,
) -> list[dict]:
    csv_paths = ensure_sendemeldungen_csv(
        base,
        csv_dir=csv_dir,
        text_cache_dir=text_cache_dir,
        force=force_csv,
    )
    records: list[dict] = []
    for csv_path in csv_paths:
        parsed = load_sendemeldung_csv(csv_path)
        print(f"  GVL CSV load: {csv_path.name} → {len(parsed):,} records", flush=True)
        records.extend(parsed)
    return records


def gvl_search_blob(rec: dict) -> str:
    parts = [
        rec.get("title", ""),
        rec.get("identification", ""),
        rec.get("performer", ""),
        rec.get("composer", ""),
        rec.get("label", ""),
        rec.get("gvlRemix", ""),
    ]
    return " ".join(p for p in parts if p)


def build_gvl_token_index(records: list[dict]) -> dict[str, list[int]]:
    token_index: dict[str, list[int]] = {}
    for idx, rec in enumerate(records):
        for tok in set(tokens_from(gvl_search_blob(rec), 2)):
            token_index.setdefault(tok, []).append(idx)
    return token_index


def load_gvl(base: Path, *, derived_dir: Path | None = None, force_sendemeldungen_csv: bool = False) -> dict:
    if not base.is_dir():
        raise FileNotFoundError(base)

    root = derived_dir or (base.parents[2] / "derived" / "cmo" / "de-gvl")
    csv_dir = root / "sendemeldungen"
    text_cache_dir = root / "pdf-text"

    records: list[dict] = []
    records.extend(load_gvl_listen(base))
    records.extend(load_gvl_produktionen(base))
    records.extend(
        load_gvl_sendemeldungen(
            base,
            csv_dir=csv_dir,
            text_cache_dir=text_cache_dir,
            force_csv=force_sendemeldungen_csv,
        )
    )

    packed = pack_source(
        source="de-gvl",
        organization="GVL",
        country="DE",
        rights_type="neighbouring",
        records=records,
    )
    packed["tokenIndex"] = build_gvl_token_index(records)
    return packed
