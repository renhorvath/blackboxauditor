"""CMO bulk file loaders → normalized record dicts."""

from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path

import openpyxl

STOP = {
    "the", "and", "feat", "ft", "featuring", "a", "an", "az", "egy", "es", "is",
    "of", "in", "on", "de", "la", "le", "les", "el", "y", "vs", "mix", "remix",
}


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    text = unicodedata.normalize("NFD", str(value))
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.lower()
    text = re.sub(r"[^\w\s]", " ", text, flags=re.UNICODE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def tokens_from(value: str | None, min_len: int = 2) -> list[str]:
    return [t for t in normalize_text(value).split() if len(t) >= min_len and t not in STOP]


def header_label(cell_value) -> str:
    if cell_value is None:
        return ""
    return str(cell_value).split("\n")[0].strip().lower()


def search_blob(rec: dict) -> str:
    parts = [
        rec.get("title", ""),
        rec.get("identification", ""),
        rec.get("performer", ""),
        rec.get("composer", ""),
        rec.get("label", ""),
    ]
    return " ".join(p for p in parts if p)


def build_identification(*, performer: str = "", composer: str = "", fallback: str = "") -> str:
    parts = [p for p in (composer, performer) if p]
    if parts:
        return " · ".join(parts)
    return fallback


def build_token_index(records: list[dict]) -> dict[str, list[int]]:
    token_index: dict[str, list[int]] = {}
    for idx, rec in enumerate(records):
        for tok in set(tokens_from(search_blob(rec), 2)):
            token_index.setdefault(tok, []).append(idx)
    return token_index


def pack_source(
    *,
    source: str,
    organization: str,
    country: str,
    rights_type: str,
    records: list[dict],
) -> dict:
    return {
        "organization": organization,
        "country": country,
        "rightsType": rights_type,
        "recordCount": len(records),
        "records": records,
        "tokenIndex": build_token_index(records),
    }


def load_akm_aume(path: Path, source: str, org: str, rights_type: str, country: str = "AT") -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        raise SystemExit(f"Empty sheet: {path}")

    col_map: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        label = header_label(cell)
        if "werknummer" in label or label == "work number":
            col_map["werknummer"] = i
        elif "werktitel" in label or label == "work title":
            col_map["werktitel"] = i
        elif "identifikation" in label or label == "identification":
            col_map["identifikation"] = i
        elif "vermerk" in label or label == "remark":
            col_map["vermerk"] = i

    records: list[dict] = []
    seen: set[str] = set()
    for row in rows_iter:
        if not row:
            continue
        werk = row[col_map["werknummer"]] if "werknummer" in col_map else None
        if werk is None:
            continue
        rec_id = str(werk).strip()
        if not rec_id or rec_id in seen:
            continue
        seen.add(rec_id)
        title = str(row[col_map["werktitel"]] if "werktitel" in col_map else "").strip()
        ident = str(row[col_map["identifikation"]] if "identifikation" in col_map else "").strip()
        remark_raw = row[col_map["vermerk"]] if "vermerk" in col_map else None
        remark = str(remark_raw).strip() if remark_raw not in (None, "") else None
        records.append(
            {
                "id": rec_id,
                "source": source,
                "title": title or "(névtelen)",
                "identification": ident,
                "remark": remark,
            }
        )
    wb.close()
    return pack_source(
        source=source, organization=org, country=country, rights_type=rights_type, records=records
    )


def _cell_str(row: tuple, col_map: dict[str, int], key: str) -> str:
    if key not in col_map or col_map[key] >= len(row):
        return ""
    val = row[col_map[key]]
    return str(val).strip() if val not in (None, "") else ""


def _map_columns(header_row: tuple) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        label = header_label(cell)
        if not label:
            continue
        if any(
            k in label
            for k in (
                "title",
                "titl",
                "titul",
                "naslov",
                "názov",
                "nazov",
                "mucim",
                "cím",
                "cime",
                "werktitel",
                "pealkiri",
                "nahravky",
                "nahrávky",
            )
        ) or (label == "werk" or label.endswith(" werk")):
            col_map.setdefault("title", i)
        if any(
            k in label
            for k in (
                "main artist",
                "performer",
                "interpret",
                "izvođač",
                "izvodac",
                "eload",
                "eloado",
                "eloadok",
                "ansambel",
                "esitaja",
                "umělci",
                "umelci",
                "vykonní",
                "vykonni",
            )
        ) or (label == "artist" or label.endswith(" artist")):
            col_map.setdefault("performer", i)
        if any(k in label for k in ("composer", "zeneszerző", "zeneszerzo", "szerző", "szerzo")):
            col_map.setdefault("composer", i)
        elif any(k in label for k in ("autor", "author")) and "publisher" not in label:
            col_map.setdefault("composer", i)
        if any(k in label for k in ("identifikation", "identification")):
            col_map.setdefault("identification", i)
        if "isrc" in label:
            col_map["isrc"] = i
        if "label" in label and "catalog" not in label:
            col_map.setdefault("label", i)
        if "gramexid" in label.replace(" ", ""):
            col_map.setdefault("external_id", i)
        if any(k in label for k in ("recording id", "werknummer", "ssz")):
            col_map.setdefault("id", i)
        elif label in ("id", "number") or label.endswith(" id"):
            col_map.setdefault("id", i)
        if any(k in label for k in ("remark", "vermerk", "note", "megjegyz")):
            col_map["remark"] = i
    return col_map


def load_flexible_xlsx(
    path: Path,
    *,
    source: str,
    id_prefix: str = "",
    sheet_tag: str | None = None,
) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    seen: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            continue
        col_map = _map_columns(header_row)
        row_num = 0
        for row in rows_iter:
            if not row:
                continue
            row_num += 1
            title = _cell_str(row, col_map, "title")
            performer = _cell_str(row, col_map, "performer")
            composer = _cell_str(row, col_map, "composer")
            ident_fallback = _cell_str(row, col_map, "identification")
            ident = build_identification(
                performer=performer, composer=composer, fallback=ident_fallback
            )
            if not title and not ident:
                continue
            rec_id = _cell_str(row, col_map, "external_id") or _cell_str(row, col_map, "id")
            if not rec_id:
                rec_id = f"{path.stem}:{sheet_name}:{row_num}"
            tag = sheet_tag or sheet_name
            dedupe = f"{id_prefix}{rec_id}:{tag}"
            if dedupe in seen:
                continue
            seen.add(dedupe)
            isrc_raw = _cell_str(row, col_map, "isrc")
            isrc = isrc_raw.upper() if isrc_raw else None
            remark_raw = _cell_str(row, col_map, "remark")
            remark = remark_raw or None
            label = _cell_str(row, col_map, "label") or None
            rec: dict = {
                "id": dedupe,
                "source": source,
                "title": title or "(névtelen)",
                "identification": ident,
                "remark": remark,
            }
            if performer:
                rec["performer"] = performer
            if composer:
                rec["composer"] = composer
            if label:
                rec["label"] = label
            if isrc:
                rec["isrc"] = isrc
            if sheet_tag or len(wb.sheetnames) > 1:
                rec["sheet"] = sheet_name
            records.append(rec)
    wb.close()
    return records


def load_csv_file(path: Path, *, source: str) -> list[dict]:
    records: list[dict] = []
    seen: set[str] = set()
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return records
        fields = [h.strip().lower() for h in reader.fieldnames]
        for row_num, row in enumerate(reader, start=2):
            norm = {k.strip().lower(): (v or "").strip() for k, v in row.items() if k}
            title = next((norm[k] for k in norm if "title" in k or "mucim" in k or "titl" in k), "")
            performer = next(
                (norm[k] for k in norm if any(x in k for x in ("artist", "performer", "interpret", "eload", "ansambel"))),
                "",
            )
            composer = next(
                (norm[k] for k in norm if any(x in k for x in ("author", "autor", "composer", "szerző", "szerzo"))),
                "",
            )
            ident = build_identification(performer=performer, composer=composer)
            if not title and not ident:
                continue
            rec_id = next((norm[k] for k in norm if k in ("id", "ssz", "work_id")), "") or f"{path.stem}:{row_num}"
            if rec_id in seen:
                continue
            seen.add(rec_id)
            isrc = next((norm[k].upper() for k in norm if "isrc" in k and norm[k]), None)
            rec: dict = {
                "id": rec_id,
                "source": source,
                "title": title or "(névtelen)",
                "identification": ident,
                "remark": None,
            }
            if performer:
                rec["performer"] = performer
            if composer:
                rec["composer"] = composer
            if isrc:
                rec["isrc"] = isrc
            records.append(rec)
    return records


def load_sena_file(path: Path, *, scope: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    seen: set[str] = set()

    for sheet_name in wb.sheetnames:
        role = "producenten" if sheet_name.lower().startswith("prod") else "muzikanten"
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        next(rows_iter, None)
        for row in rows_iter:
            if not row or len(row) < 3:
                continue
            rec_id = str(row[0]).strip() if row[0] is not None else ""
            if not rec_id:
                continue
            dedupe_key = f"{rec_id}:{role}:{scope}"
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            artist = str(row[1]).strip() if row[1] is not None else ""
            title = str(row[2]).strip() if row[2] is not None else ""
            version = str(row[3]).strip() if len(row) > 3 and row[3] not in (None, "") else None
            isrc_raw = row[4] if len(row) > 4 else None
            isrc = str(isrc_raw).strip().upper() if isrc_raw not in (None, "") else None
            full_title = f"{title} ({version})" if version else title
            records.append(
                {
                    "id": dedupe_key,
                    "source": "nl-sena",
                    "title": full_title or "(névtelen)",
                    "identification": artist,
                    "performer": artist or None,
                    "remark": None,
                    "senaRole": role,
                    "senaScope": scope,
                    "isrc": isrc,
                }
            )
    wb.close()
    return records


def load_sena(dir_path: Path) -> dict:
    files = [
        (dir_path / "ongeclaimd-nederland.xlsx", "nederland"),
        (dir_path / "ongeclaimd-buitenland.xlsx", "buitenland"),
    ]
    records: list[dict] = []
    for path, scope in files:
        if not path.is_file():
            raise SystemExit(f"Missing SENA file: {path}")
        records.extend(load_sena_file(path, scope=scope))
    return pack_source(
        source="nl-sena",
        organization="SENA",
        country="NL",
        rights_type="neighbouring",
        records=records,
    )


def load_dir_xlsx(dir_path: Path, *, source: str, org: str, country: str, rights_type: str) -> dict:
    paths = sorted(dir_path.glob("*.xlsx")) + sorted(dir_path.glob("*.xls"))
    if not paths:
        raise FileNotFoundError(dir_path)
    records: list[dict] = []
    for path in paths:
        records.extend(load_flexible_xlsx(path, source=source, id_prefix=f"{path.name}:"))
    return pack_source(
        source=source, organization=org, country=country, rights_type=rights_type, records=records
    )


def load_dir_csv(dir_path: Path, *, source: str, org: str, country: str, rights_type: str) -> dict:
    paths = sorted(dir_path.glob("*.csv"))
    if not paths:
        raise FileNotFoundError(dir_path)
    records: list[dict] = []
    for path in paths:
        records.extend(load_csv_file(path, source=source))
    return pack_source(
        source=source, organization=org, country=country, rights_type=rights_type, records=records
    )
