#!/usr/bin/env python3
"""Build searchable JSON indexes from AKM, AUME, and SENA xlsx sources."""

from __future__ import annotations

import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUT_PATH = PROJECT_ROOT / "data" / "cmo-index.json"

STOP = {
    "the", "and", "feat", "ft", "featuring", "a", "an", "az", "egy", "es", "is",
    "of", "in", "on", "de", "la", "le", "les", "el", "y", "vs", "mix", "remix",
}


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    text = unicodedata.normalize("NFD", str(value))
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.lower()
    text = re.sub(r"[^\w\s]", " ", text, flags=re.UNICODE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def tokens_from(value: str | None, min_len: int = 2) -> list[str]:
    return [t for t in normalize_text(value).split() if len(t) >= min_len and t not in STOP]


def header_label(cell_value) -> str:
    if cell_value is None:
        return ""
    return str(cell_value).split("\n")[0].strip().lower()


def load_akm_aume(path: Path, source: str, org: str, rights_type: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        raise SystemExit(f"Empty sheet: {path}")

    col_map: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        label = header_label(cell)
        if "werknummer" in label or label == "work number":
            col_map["werknummer"] = i
        elif "werktitel" in label or label == "work title":
            col_map["werktitel"] = i
        elif "identifikation" in label or label == "identification":
            col_map["identifikation"] = i
        elif "vermerk" in label or label == "remark":
            col_map["vermerk"] = i

    records = []
    seen: set[str] = set()
    for row in rows_iter:
        if not row:
            continue
        werk = row[col_map["werknummer"]] if "werknummer" in col_map else None
        if werk is None:
            continue
        rec_id = str(werk).strip()
        if not rec_id or rec_id in seen:
            continue
        seen.add(rec_id)
        title = str(row[col_map["werktitel"]] if "werktitel" in col_map else "").strip()
        ident = str(row[col_map["identifikation"]] if "identifikation" in col_map else "").strip()
        remark_raw = row[col_map["vermerk"]] if "vermerk" in col_map else None
        remark = str(remark_raw).strip() if remark_raw not in (None, "") else None
        records.append(
            {
                "id": rec_id,
                "source": source,
                "title": title or "(névtelen)",
                "identification": ident,
                "remark": remark,
            }
        )
    wb.close()

    token_index = build_token_index(records)
    return {
        "organization": org,
        "country": "AT",
        "rightsType": rights_type,
        "recordCount": len(records),
        "records": records,
        "tokenIndex": token_index,
    }


def load_sena(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []
    seen: set[str] = set()

    for sheet_name in wb.sheetnames:
        role = "producenten" if sheet_name.lower().startswith("prod") else "muzikanten"
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        next(rows_iter, None)  # skip header
        for row in rows_iter:
            if not row or len(row) < 3:
                continue
            rec_id = str(row[0]).strip() if row[0] is not None else ""
            if not rec_id:
                continue
            dedupe_key = f"{rec_id}:{role}"
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            artist = str(row[1]).strip() if row[1] is not None else ""
            title = str(row[2]).strip() if row[2] is not None else ""
            version = str(row[3]).strip() if len(row) > 3 and row[3] not in (None, "") else None
            isrc_raw = row[4] if len(row) > 4 else None
            isrc = str(isrc_raw).strip().upper() if isrc_raw not in (None, "") else None
            full_title = title
            if version:
                full_title = f"{title} ({version})"
            records.append(
                {
                    "id": dedupe_key,
                    "source": "nl-sena",
                    "title": full_title or "(névtelen)",
                    "identification": artist,
                    "remark": None,
                    "senaRole": role,
                    "isrc": isrc,
                }
            )
    wb.close()

    token_index = build_token_index(records)
    return {
        "organization": "SENA",
        "country": "NL",
        "rightsType": "neighbouring",
        "recordCount": len(records),
        "records": records,
        "tokenIndex": token_index,
    }


def build_token_index(records: list[dict]) -> dict[str, list[int]]:
    token_index: dict[str, list[int]] = {}
    for idx, rec in enumerate(records):
        blob = f"{rec['title']} {rec['identification']}"
        seen_tokens = set(tokens_from(blob, 2))
        for tok in seen_tokens:
            token_index.setdefault(tok, []).append(idx)
    return token_index


def main() -> None:
    raw = PROJECT_ROOT / "raw" / "cmo"
    sources = {
        "at-akm": load_akm_aume(
            raw / "at-akm" / "Anfrageliste-AKM-allgemein.xlsx",
            "at-akm",
            "AKM",
            "musical_work",
        ),
        "at-aume": load_akm_aume(
            raw / "at-aume" / "Anfrageliste-aume-allgemein.xlsx",
            "at-aume",
            "Austro-Mechana",
            "mechanical",
        ),
        "nl-sena": load_sena(raw / "nl-sena" / "ongeclaimd-buitenland.xlsx"),
    }

    payload = {
        "version": 1,
        "builtAt": datetime.now(timezone.utc).isoformat(),
        "sources": sources,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload), encoding="utf-8")
    total = sum(s["recordCount"] for s in sources.values())
    print(f"Wrote {OUT_PATH} ({total:,} records across {len(sources)} sources)")
    for sid, meta in sources.items():
        print(f"  {sid}: {meta['recordCount']:,} ({meta['organization']})")


if __name__ == "__main__":
    main()
